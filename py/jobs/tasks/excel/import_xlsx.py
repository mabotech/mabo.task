# -*- coding: utf-8 -*-

"""
upload and save, post jsonrpc msg?
show in grid and editable
edit and save

save to db and return json

json or sql?

"""
import json

import datetime

import openpyxl

#from openpyxl.cell import get_column_letter


def save():
    """ http / requests
    or pg / plv8 function? 
    """
    
    pass

def xlsx(filename):    
    """   """
    
    xls_book = openpyxl.load_workbook(filename=filename)

    sheet_names = xls_book.get_sheet_names()

    xls_sheet = xls_book.get_sheet_by_name(sheet_names[0])

    max_row = xls_sheet.max_row
    
    headers = ["productline","projectname","taskno","sampleno","vin","name","section","planstart","planend"]
    
    for rownum in range(1, max_row):
        # print rownum
        
        row = {}
        
        line = []
        
        
        for colnum in range(9):
            
            
            
            val = xls_sheet.cell(row=rownum + 1, column=colnum + 1).value
            
            
            if val is None:
                break

            val_type = type(val)

            if isinstance(val, unicode):
                x =val.encode("utf8")
            
            elif isinstance(val, datetime.datetime):            
                x = val.isoformat()
            
            else:
                
                x = val
                
            colname = headers[colnum]
            row[colname] = x
            line.append(x)
                
        if colnum == 8:
            print json.dumps(row)   
            print line
            print("\n")
            

def xls(filename):
    
    print filename.encode("utf8")


if __name__ == "__main__":
    
    filename  = "weekly_plan_template.xlsx"
    
    xlsx(filename)
    
    filename = u"下周试验计划导入模板.xls"
    
    xls(filename)    