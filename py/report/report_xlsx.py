# -*- coding: utf-8 -*-

"""
fetch data from PostgreSQL
get JSON result?

"""
from time import strftime, localtime

import openpyxl


class ReportXlsx(object):

    def __init__(self, filename):
        """ init """
        self.xls_book = openpyxl.load_workbook(filename=filename)

        self.sheet_names = self.xls_book.get_sheet_names()
        
        # pg, json?

    def sheet1(self):
        """ sheet 1 """
        xls_sheet = self.xls_book.get_sheet_by_name(self.sheet_names[0])

        xls_sheet["C2"] = 0.6226
        xls_sheet["D2"] = 0.7277
        xls_sheet["E2"] = 0.782

        xls_sheet["E10"] = strftime("%Y-%m-%d", localtime())

    def sheet2(self):
        """ sheet 2 """

        xls_sheet = self.xls_book.get_sheet_by_name(self.sheet_names[1])

        xls_sheet["A1"] = "2014年12月试验中心主要设备（单班）利用率、完好率报表"
        xls_sheet["J3"] = 0.91

    def sheet3(self):
        """ sheet 3 """

        xls_sheet = self.xls_book.get_sheet_by_name(self.sheet_names[2])

        # print dir(xls_sheet)#.decode("utf8")

        # print xls_sheet.title.encode("utf8")

        # print xls_sheet.max_column
        #xls_sheet["AC4"] = 62.6

        xls_sheet.cell(row=4, column=28).value = 0.912
        xls_sheet.cell(row=5, column=28).value = 0.863
        xls_sheet.cell(row=4, column=29).value = 1
        xls_sheet.cell(row=5, column=29).value = 1

    def generate(self, filename):
        """ generate report """

        self.sheet1()
        self.sheet2()
        self.sheet3()

        self.xls_book.save(filename)


def main():

    template = 'report_201412.xlsx'

    xlsx = ReportXlsx(template)

    filename = "report_201412-3.xlsx"

    xlsx.generate(filename)

if __name__ == "__main__":

    main()
