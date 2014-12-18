# -*- coding: utf-8 -*-


import openpyxl


def sheet1(xls_book, sheet_names):
    xls_sheet = xls_book.get_sheet_by_name(sheet_names[0])

    xls_sheet["C2"] = 0.6226
    xls_sheet["D2"] = 0.7277
    xls_sheet["E2"] = 0.782
    
    
def sheet2(xls_book, sheet_names):
    xls_sheet = xls_book.get_sheet_by_name(sheet_names[1])

    xls_sheet["A1"] = "2014年12月试验中心主要设备（单班）利用率、完好率报表"
    xls_sheet["J3"] = 0.91
    
    
def sheet3(xls_book, sheet_names):
    xls_sheet = xls_book.get_sheet_by_name(sheet_names[2])

    #print dir(xls_sheet)#.decode("utf8")

    #print xls_sheet.title.encode("utf8")

    #print xls_sheet.max_column
    #xls_sheet["AC4"] = 62.6

    xls_sheet.cell(row = 4, column = 28).value = 0.912
    xls_sheet.cell(row = 5, column = 28).value = 0.863

    xls_sheet.cell(row = 4, column = 29).value = 1
    xls_sheet.cell(row = 5, column = 29).value = 1
    

def main():
    
    xls_book = openpyxl.load_workbook(filename=r'report_201412.xlsx')

    sheet_names = xls_book.get_sheet_names()

    sheet1(xls_book,sheet_names)
    sheet2(xls_book,sheet_names)
    sheet3(xls_book,sheet_names)
    
    """
    xls_sheet = xls_book.get_sheet_by_name(sheet_names[0])

    xls_sheet["C2"] = 0.52
    xls_sheet["D2"] = 0.756
    xls_sheet["E2"] = 0.782
    """
    




    excel_file = "report_201412-2.xlsx"

    xls_book.save(excel_file)
    
if __name__ == "__main__":

    main()