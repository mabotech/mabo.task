# -*- coding: utf-8 -*-

"""
fetch data from PostgreSQL
"""

import os
#import calendar
import json
import shutil
import datetime
from time import strftime, localtime

import toml
import requests
import openpyxl



conf_fn = os.sep.join(
    [os.path.split(os.path.realpath(__file__))[0], "report.toml"])

# print conf_fn

with open(conf_fn) as conf_fh:

    cfg = toml.loads(conf_fh.read())

    conf = cfg["app"]


class ReportXlsx(object):

    def __init__(self, filename):
        """ init """
        self.xls_book = openpyxl.load_workbook(filename=filename)
        
        (self.year,self.month) = get_prev_month()

        self.sheet_names = self.xls_book.get_sheet_names()
        
        # pg, json?

    def sheet1(self, data):
        """ sheet 1 """
        
        xls_sheet = self.xls_book.get_sheet_by_name(self.sheet_names[0])

        i = 0
        for row in data:
            #print row.equipid, row.plantime, row.worktime
            xls_sheet.cell(row=2+i, column=2).value = row["equipname"]
            xls_sheet.cell(row=2+i, column=3).value = row["equipid"]
            xls_sheet.cell(row=2+i, column=4).value = row["plantime"]
            xls_sheet.cell(row=2+i, column=5).value = row["worktime"]
            i = i +1
        
        xls_sheet["F15"] = "生成时间:%s" %(strftime("%Y-%m-%d", localtime()) )

    def sheet2(self, data):
        """ sheet 2 """

        xls_sheet = self.xls_book.get_sheet_by_name(self.sheet_names[1])

        xls_sheet["A1"] = "%s年%s月试验中心主要设备（单班）利用率、完好率报表" %(self.year, self.month)
        i = 0
        for row in data:
            xls_sheet.cell(row=3+i, column=3).value = row["equipname"]
            xls_sheet.cell(row=3+i, column=4).value = row["equipid"]
            xls_sheet.cell(row=3+i, column=7).value = row["plantime"]
            xls_sheet.cell(row=3+i, column=6).value = row["worktime"]
            xls_sheet.cell(row=3+i, column=9).value = row["downtime"]
            i = i +1

    def sheet3(self, data):
        """ sheet 3 """

        xls_sheet = self.xls_book.get_sheet_by_name(self.sheet_names[2])

        xls_sheet["A1"] = "%s年度试验中心主要设备（单班）利用率、完好率报表" %(self.year)
        
        # 6， 8， 10，12，14，16
        # 18，20，22，24，26，28
        
        i = 0
        
        v1 = 0
        v2 = 0
        for row in data:

            xls_sheet.cell(row=4+i, column=3).value =  row["equipname"]
            xls_sheet.cell(row=4+i, column=4).value = row["equipid"]            
            xls_sheet.cell(row=4+i, column=4+self.month*2).value =  row["kpi1"]
            xls_sheet.cell(row=4+i, column=5+self.month*2).value = row["kpi2"]
            
            v1 = v1 +row["kpi1"]
            v2 = v2 +row["kpi2"]
            
            i = i +1
        
        xls_sheet.cell(row=10, column=4+self.month*2).value = v1 / i
        xls_sheet.cell(row=10, column=5+self.month*2).value = v2 /i


    def generate(self, filename):
        """ generate report """
        
        method = conf["METHOD"]
        
        data = self.get_data(method)
        
        self.sheet1(data)
        self.sheet2(data)
        self.sheet3(data)

        self.xls_book.save(filename)

    def get_data(self, method):
        
       
        year_month = "%s-%02d" %(self.year, self.month)
        
        URL = conf["JSONRPC"]

        payload = {
            "jsonrpc": "2.0",
            "id": "r2",
            "method": "call",
            "params": {
                "method": method,
                "month":year_month,
                "context": {
                    "user": "mt",
                    "languageid": "1033",
                    "sessionid": "123"}}}

        HEADERS = {
            'content-type': 'application/json',
            'accept': 'json',
            'User-Agent': 'mabo'}

        payload = json.dumps(payload)

        resp = requests.post(URL, data=payload, headers=HEADERS)

        s = resp.text  # .encode("utf8")
        
        #print s.encode("utf-8")
        
        v = json.loads(s)
        
        return v["retuning"]


def get_prev_month():
    
    today = datetime.date.today()
    first = datetime.date(day=1, month=today.month, year=today.year)
    lastMonth = first - datetime.timedelta(days=1)
    
    return (int(lastMonth.strftime("%Y")), int(lastMonth.strftime("%m") ) )


def main():

    template = conf["template"]
    xlsx = ReportXlsx(template)
    output = conf["output"] 
    
    xlsx.generate(output)
    
    ## copy file for rolling file
    ## TODO: reset when Feb.
    try:
        shutil.copyfile(output, template)
    except Exception as ex:
        pass    
    
if __name__ == "__main__":

    main()
    