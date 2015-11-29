# -*- coding: utf-8 -*-


import openpyxl


xls_book = openpyxl.load_workbook(filename=r'stack37.xlsx')

sheet_names = xls_book.get_sheet_names()


xls_sheet = xls_book.get_sheet_by_name(sheet_names[0])

xls_sheet["C2"] = 0.52
xls_sheet["D2"] = 0.756
xls_sheet["E2"] = 0.782

xls_sheet = xls_book.get_sheet_by_name(sheet_names[1])

xls_sheet["A1"] = "2014年12月试验中心主要设备（单班）利用率、完好率报表"

xls_sheet = xls_book.get_sheet_by_name(sheet_names[2])

print dir(xls_sheet)#.decode("utf8")

print xls_sheet.title.encode("utf8")

print xls_sheet.max_column
#xls_sheet["AC4"] = 62.6

xls_sheet.cell(row = 4, column = 28).value = 0.912
xls_sheet.cell(row = 5, column = 28).value = 0.863

xls_sheet.cell(row = 4, column = 29).value = 1
xls_sheet.cell(row = 5, column = 29).value = 1

xls_book.save('stack37_3.xlsx')